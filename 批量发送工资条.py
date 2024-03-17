#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@文件        :批量发送工资条.py
@说明        :
@时间        :2024/03/08 14:34:25
@作者        :xnp2010
@版本        :1.0
'''

import openpyxl  # pip install openpyxl
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# 设置发件人、收件人和邮件内容
sender_email = "nianping.xia@insolu.cn"
password = "PCgRaG3HteehUJLN"

subject = "工资条"


wb = openpyxl.load_workbook('工资表.xlsx', data_only=True)
sheet = wb['Sheet1']

month = sheet['G2'].value

try:
    server = smtplib.SMTP("smtp.exmail.qq.com", 587)  # 设置SMTP服务器地址和端口号
    server.starttls()  # 开启TLS加密
    server.login(sender_email, password)  # 登录邮箱

    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row-2, min_col=2):
        name, email, occupation, attendance, basic_salary, bonus, deduction, final_salary = row[
            0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value, row[7].value

        html_body = f"""
<!DOCTYPE html>
<html lang="en">

<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Document</title>
    <style>
    table, th, td {{
    border: 1px solid black;
    border-collapse: collapse;
    }}
</style>
</head>

<body>
    <h3>Dear {name}:</h3>
    <p>辛苦了！请查收你{month}的工资条：</p>
    <table>
        <tr>
            <th>姓名</th>
            <th>邮箱</th>
            <th>职位</th>
            <th>实出勤天数</th>
            <th>基本工资</th>
            <th>奖  金</th>
            <th>应扣款项</th>
            <th>实发金额</th>
        </tr>
        <tr>
            <td>{name}</td>
            <td>{email}</td>
            <td>{occupation}</td>
            <td>{attendance}</td>
            <td>{basic_salary}</td>
            <td>{bonus}</td>
            <td>{deduction}</td>
            <td>{final_salary}</td>
        </tr>
    </table>


</body>

</html>
"""
        # 创建邮件对象
        message = MIMEText(html_body, "html", "utf-8")
        message["From"] = sender_email
        message["To"] = email
        message["Subject"] = subject

        server.sendmail(sender_email, email, message.as_string())  # 发送邮件

        print(f"{name}工资条已发送到{email}!")
except Exception as e:
    print("Error:", e)
finally:
    server.quit()  # 关闭连接

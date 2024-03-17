#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@文件        :批量发送工资条 copy.py
@说明        :
@时间        :2024/03/09 16:29:29
@作者        :xnp2010
@版本        :1.0
'''

from openpyxl.styles import Font, colors, Alignment
from openpyxl import Workbook, load_workbook
import datetime

# 创建excel
# 实例化
wb = Workbook()
# 获取当前active的sheet
ws = wb.active
print(ws)
print(type(ws))
print(ws.title)
ws.title = 'salary'

# 写数据
# 方式一
ws['C5'] = 'C5的数据'

# 方式二：附加在最下方的空行
ws.append([1, 2, 3])


# 方式三：python类型自动转换
ws['A1'] = datetime.datetime.now().strftime('%Y-%m-%d')

# 选择sheet页
chose_sheet = wb['salary']
chose_sheet2 = wb.get_sheet_by_name('salary')
print(wb.get_sheet_names())  # 打印所有的sheet
chose_sheet3 = wb.worksheets[0]  # 获得第1个sheet

# 删除sheet页
# 方式一：
wb.remove(chose_sheet)
# 方式二：
del wb['salary']


wb.save('创建的工资表.xlsx')

# 打开已有excel
wb2 = load_workbook('工资表.xlsx')

sheet = wb2.get_sheet_by_name('sheet1')
# 按行遍历
for row in sheet:
    for cell in row:
        print(cell.value, end=',')
    print()

# 按列遍历
for column in sheet.columns:
    for cell in column:
        print(cell.value, end=',')
    print()

# 遍历指定行&列
for row in sheet.iter_rows(min_row=3, max_row=5, max_col=5):  # 遍历3-5行，1-5列
    for cell in row:
        print(cell.value, end=',')
    print()

for col in sheet.iter_cols(min_col=1, max_col=5, min_row=3):  # 遍历1-5列，3-最大行
    for cell in col:
        print(cell.value, end=',')
    print()

# 设置单元格格式
bold_itatic_24_font = Font(
    name='等线', size=24, italic=True, color=colors.RED, bold=True)  # 声明样式
sheet['A1'].font = bold_itatic_24_font  # 设置单元格样式
sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')  # 居中
sheet.row_dimensions[2].height = 40  # 设置第2行的行高
sheet.column_dimensions['C'].width = 30  # 设置C列列宽
